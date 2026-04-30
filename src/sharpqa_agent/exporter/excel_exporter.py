"""Excel exporter — produces formatted .xlsx files with lead data and email drafts."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from sharpqa_agent.core.database import (
    get_contacts_for_lead,
    get_drafts,
    get_findings_for_lead,
    get_leads,
    get_tech_stack_for_lead,
)
from sharpqa_agent.core.logging_setup import get_logger

logger = get_logger(__name__)

# Column definitions matching the design spec
COLUMNS = [
    "Company Name",
    "Website",
    "Primary Email",
    "Primary Contact",
    "Phone",
    "LinkedIn",
    "Twitter",
    "Top Findings Summary",
    "Improvements Summary",
    "Estimated Business Impact",
    "Draft Subject",
    "Draft Body",
    "Priority Score",
    "Status",
]


async def export_leads_to_excel(
    settings,
    lead_ids: list[str] | None = None,
    status_filter: str | None = None,
) -> Path:
    """Export leads with enrichment data and drafts to an Excel file.

    Args:
        settings: Application settings (for db path and export dir).
        lead_ids: Optional specific lead IDs to export. If None, exports all.
        status_filter: Optional status filter for leads.

    Returns:
        Path to the generated Excel file.
    """
    db_path = settings.sqlite_db_path
    exports_dir = Path(settings.exports_dir)
    exports_dir.mkdir(parents=True, exist_ok=True)

    # Fetch leads
    if lead_ids:
        from sharpqa_agent.core.database import get_lead_by_id
        leads = []
        for lid in lead_ids:
            lead = await get_lead_by_id(db_path, lid)
            if lead:
                leads.append(lead)
    else:
        leads = await get_leads(db_path, status=status_filter, limit=1000)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "SharpQA Leads"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")

    # Write headers
    for col_index, header in enumerate(COLUMNS, 1):
        cell = sheet.cell(row=1, column=col_index, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Freeze top row
    sheet.freeze_panes = "A2"

    # Write data rows
    for row_index, lead in enumerate(leads, 2):
        # Get related data
        contacts = await get_contacts_for_lead(db_path, lead.lead_id)
        findings = await get_findings_for_lead(db_path, lead.lead_id)
        tech_stack = await get_tech_stack_for_lead(db_path, lead.lead_id)
        drafts = await get_drafts(db_path, lead_id=lead.lead_id, limit=1)

        primary_contact = next((c for c in contacts if c.is_primary_contact), contacts[0] if contacts else None)
        draft = drafts[0] if drafts else None

        # Summarize findings
        top_findings = "; ".join(
            f"[{f.severity_level.value}] {f.finding_title}"
            for f in findings[:3]
        ) if findings else "No findings"

        improvements = "; ".join(
            f.finding_description[:80] for f in findings[:3] if f.finding_description
        ) if findings else ""

        impact = findings[0].business_impact if findings and findings[0].business_impact else ""

        row_data = [
            lead.company_name,
            lead.website_url,
            primary_contact.email_address if primary_contact else "",
            primary_contact.full_name if primary_contact else "",
            "",  # Phone — not typically scraped
            primary_contact.linkedin_url if primary_contact else "",
            primary_contact.twitter_handle if primary_contact else "",
            top_findings,
            improvements[:300],
            impact[:300],
            draft.subject_line if draft else "",
            (draft.human_edited_body or draft.email_body) if draft else "",
            round(lead.priority_score, 2),
            lead.lead_status.value if hasattr(lead.lead_status, 'value') else lead.lead_status,
        ]

        for col_index, value in enumerate(row_data, 1):
            cell = sheet.cell(row=row_index, column=col_index, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Auto-size columns
    for col_index in range(1, len(COLUMNS) + 1):
        column_letter = get_column_letter(col_index)
        max_width = len(COLUMNS[col_index - 1])
        for row in range(2, min(len(leads) + 2, 20)):  # Sample first 20 rows
            cell_value = str(sheet.cell(row=row, column=col_index).value or "")
            max_width = max(max_width, min(len(cell_value), 50))
        sheet.column_dimensions[column_letter].width = max_width + 2

    # Generate filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = exports_dir / f"SharpQA_leads_{timestamp}.xlsx"

    workbook.save(str(output_path))
    logger.info("excel_exported", path=str(output_path), leads=len(leads))

    return output_path
